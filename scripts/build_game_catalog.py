#!/usr/bin/env python3
from __future__ import annotations

import json
import hashlib
import re
import shutil
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
ASSETS_DIR = ROOT_DIR / "assets"
ICONS_DIR = ASSETS_DIR / "icons"
ICONS_EN_DIR = ASSETS_DIR / "icons_en"
SOURCE_XLSX = ASSETS_DIR / "configs.xlsx"
OUTPUT_XLSX = ASSETS_DIR / "configs_game.xlsx"
OUTPUT_JS = ROOT_DIR / "src" / "game_data.generated.js"

PACK_KEY_MAP = {
    "1乌龟兽群": {"key": "pack1", "name_en": "Pack 1"},
    "2濒危兽群": {"key": "pack2", "name_en": "Pack 2"},
    "3独角兽兽群": {"key": "pack3", "name_en": "Pack 3"},
    "4金毛兽群": {"key": "pack4", "name_en": "Pack 4"},
    "5海星兽群": {"key": "pack5", "name_en": "Pack 5"},
    "6幼犬兽群": {"key": "pack6", "name_en": "Pack 6"},
}

# Keep existing gameplay behavior for the current implemented roster.
LEGACY_PET_KIND_MAP = {
    "蚂蚁": {
        "kind": "ant",
        "name_en": "Ant",
        "color": "#d98f4e",
        "ability_key": "faint_buff_random_ally",
        "attack": 2,
        "health": 1,
    },
    "鱼": {
        "kind": "fish",
        "name_en": "Fish",
        "color": "#72b8ea",
        "ability_key": "level_buff_team",
        "attack": 2,
        "health": 3,
    },
    "河狸": {
        "kind": "beaver",
        "name_en": "Beaver",
        "color": "#ab7d61",
        "ability_key": "sell_buff_attack",
        "attack": 2,
        "health": 2,
    },
    "蛐蛐": {
        "kind": "cricket",
        "name_en": "Cricket",
        "color": "#8ac95f",
        "ability_key": "faint_summon_zombie",
        "attack": 1,
        "health": 2,
    },
    "蚊子": {
        "kind": "mosquito",
        "name_en": "Mosquito",
        "color": "#8a7ef2",
        "ability_key": "start_ping_enemy",
        "attack": 2,
        "health": 2,
    },
    "水獭": {
        "kind": "otter",
        "name_en": "Otter",
        "color": "#8f9fc2",
        "ability_key": "buy_buff_random_shop_pet",
        "attack": 1,
        "health": 2,
    },
    "天鹅": {
        "kind": "swan",
        "name_en": "Swan",
        "color": "#e3f3ff",
        "ability_key": "start_turn_gain_gold",
        "attack": 1,
        "health": 3,
    },
    "马": {
        "kind": "horse",
        "name_en": "Horse",
        "color": "#af8868",
        "ability_key": "friend_summoned_attack",
        "attack": 2,
        "health": 1,
    },
    "火烈鸟": {
        "kind": "flamingo",
        "name_en": "Flamingo",
        "color": "#f08da8",
        "ability_key": "faint_buff_rear",
        "attack": 3,
        "health": 2,
    },
    "骆驼": {
        "kind": "camel",
        "name_en": "Camel",
        "color": "#d1ba75",
        "ability_key": "hurt_buff_rear",
        "attack": 2,
        "health": 5,
    },
    "袋鼠": {
        "kind": "kangaroo",
        "name_en": "Kangaroo",
        "color": "#c58c42",
        "ability_key": "behind_attack_gain",
        "attack": 1,
        "health": 3,
    },
    "长颈鹿": {
        "kind": "giraffe",
        "name_en": "Giraffe",
        "color": "#e8b85e",
        "ability_key": "end_turn_buff_friend_ahead",
        "attack": 2,
        "health": 4,
    },
    "家兔": {
        "kind": "rabbit",
        "name_en": "Rabbit",
        "color": "#f5f0f7",
        "ability_key": "friend_eat_bonus_health",
        "attack": 3,
        "health": 2,
    },
    "孔雀": {
        "kind": "peacock",
        "name_en": "Peacock",
        "color": "#76b7c2",
        "ability_key": "hurt_gain_attack",
        "attack": 2,
        "health": 5,
    },
    "渡渡鸟": {
        "kind": "dodo",
        "name_en": "Dodo",
        "color": "#d7ad7b",
        "ability_key": "start_battle_buff_friend_ahead_attack",
        "attack": 2,
        "health": 3,
    },
    "企鹅": {
        "kind": "penguin",
        "name_en": "Penguin",
        "color": "#5d7d9f",
        "ability_key": "end_turn_buff_level2_friends",
        "attack": 1,
        "health": 2,
    },
    "乌龟": {
        "kind": "turtle",
        "name_en": "Turtle",
        "color": "#6aaa77",
        "ability_key": "faint_give_melon_friend_behind",
        "attack": 1,
        "health": 2,
    },
}

LEGACY_FOOD_KIND_MAP = {
    "苹果": {
        "kind": "apple",
        "name_en": "Apple",
        "color": "#e55e57",
        "effect_key": "stat_1_1",
    },
    "蜂蜜": {
        "kind": "honey",
        "name_en": "Honey",
        "color": "#e0aa35",
        "effect_key": "summon_bee",
    },
    "肉骨头": {
        "kind": "meat",
        "name_en": "Meat",
        "color": "#d66a4f",
        "effect_key": "perk_meat",
    },
    "大蒜": {
        "kind": "garlic",
        "name_en": "Garlic",
        "color": "#b9b3a6",
        "effect_key": "perk_garlic",
    },
    "鸭梨": {
        "kind": "pear",
        "name_en": "Pear",
        "color": "#8dc95f",
        "effect_key": "stat_2_2",
    },
    "碗装沙拉": {
        "kind": "salad",
        "name_en": "Salad",
        "color": "#62c26e",
        "effect_key": "team_random_buff_1_1",
    },
    "纸杯蛋糕": {
        "kind": "cupcake",
        "name_en": "Cupcake",
        "color": "#f29ec0",
        "effect_key": "temp_3_3",
    },
    "食物罐头": {
        "kind": "canned_food",
        "name_en": "Can",
        "color": "#8ea7bb",
        "effect_key": "shop_buff_2_1",
    },
    "蜜瓜": {
        "kind": "melon",
        "name_en": "Melon",
        "color": "#8fdc78",
        "effect_key": "perk_melon",
    },
    "巧克力": {
        "kind": "chocolate",
        "name_en": "Chocolate",
        "color": "#9f7a58",
        "effect_key": "exp_1",
    },
}

CATALOG_COLUMNS = [
    "id",
    "pack_key",
    "pack_name_zh",
    "pack_name_en",
    "type",
    "name_zh",
    "name_en",
    "tier",
    "round_unlock",
    "icon_src_rel",
    "icon_alias_rel",
    "hint_zh_raw",
    "hint_zh_clean",
    "hint_en",
    "attack",
    "health",
    "ability_key",
    "effect_key",
    "impl_status",
    "source_sheet",
    "source_row",
]


def normalize_name(name: str) -> str:
    text = unicodedata.normalize("NFKC", str(name or ""))
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"\d+$", "", text)
    return text


def clean_hint(raw: str) -> str:
    if not raw:
        return ""
    text = unicodedata.normalize("NFKC", str(raw))
    text = text.replace("\r", "\n")
    lines: list[str] = []
    for raw_line in text.split("\n"):
        line = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff\s+\-*/%:，。！？、,.()（）【】\[\]<>]", " ", raw_line)
        line = re.sub(r"\s+", " ", line).strip()
        if len(line) >= 2:
            lines.append(line)
        if len(lines) >= 3:
            break
    return "\n".join(lines)


def fnv1a(value: str) -> int:
    h = 0x811C9DC5
    for b in value.encode("utf-8"):
        h ^= b
        h = (h * 0x01000193) % (1 << 32)
    return h


def clamp(value: int, low: int, high: int) -> int:
    return max(low, min(high, value))


def auto_pet_stats(item_id: str, tier: int) -> tuple[int, int]:
    seed = fnv1a(item_id)
    attack = clamp(tier + 1 + ((seed % 3) - 1), 1, 12)
    health = clamp(tier + 2 + (((seed >> 3) % 3) - 1), 1, 15)
    return attack, health


def placeholder_food_buff(tier: int) -> int:
    if tier <= 2:
        return 1
    if tier <= 4:
        return 2
    return 3


def color_from_id(item_id: str, is_food: bool) -> str:
    seed = fnv1a(item_id)
    # Pastel HSL -> RGB
    hue = seed % 360
    sat = 46 if is_food else 42
    light = 73 if is_food else 69
    c = (1 - abs(2 * light / 100 - 1)) * sat / 100
    x = c * (1 - abs((hue / 60) % 2 - 1))
    m = light / 100 - c / 2
    if 0 <= hue < 60:
        r1, g1, b1 = c, x, 0
    elif 60 <= hue < 120:
        r1, g1, b1 = x, c, 0
    elif 120 <= hue < 180:
        r1, g1, b1 = 0, c, x
    elif 180 <= hue < 240:
        r1, g1, b1 = 0, x, c
    elif 240 <= hue < 300:
        r1, g1, b1 = x, 0, c
    else:
        r1, g1, b1 = c, 0, x
    r = int(round((r1 + m) * 255))
    g = int(round((g1 + m) * 255))
    b = int(round((b1 + m) * 255))
    return f"#{r:02x}{g:02x}{b:02x}"


@dataclass
class RawRow:
    pack_name_zh: str
    tier: int
    round_unlock: int
    type_zh: str
    icon_src_rel: str
    name_raw: str
    hint_raw: str
    ocr_score: float
    source_sheet: str
    source_row: int


def parse_row(ws: Any, row_idx: int) -> RawRow | None:
    pack_name_zh = str(ws[f"B{row_idx}"].value or "").strip()
    tier_val = ws[f"D{row_idx}"].value
    round_val = ws[f"E{row_idx}"].value
    type_zh = str(ws[f"F{row_idx}"].value or "").strip()
    icon_src_rel = str(ws[f"H{row_idx}"].value or "").strip().replace("\\", "/")
    name_raw = str(ws[f"I{row_idx}"].value or "").strip()
    hint_raw = str(ws[f"J{row_idx}"].value or "").strip()
    score_val = ws[f"M{row_idx}"].value

    if not (pack_name_zh and type_zh and icon_src_rel and name_raw):
        return None
    if pack_name_zh not in PACK_KEY_MAP:
        return None
    if type_zh not in ("动物", "食物"):
        return None

    tier = int(tier_val) if tier_val is not None else 1
    round_unlock = int(round_val) if round_val is not None else 1
    try:
        ocr_score = float(score_val) if score_val is not None else 0.0
    except Exception:
        ocr_score = 0.0

    return RawRow(
        pack_name_zh=pack_name_zh,
        tier=tier,
        round_unlock=round_unlock,
        type_zh=type_zh,
        icon_src_rel=icon_src_rel,
        name_raw=name_raw,
        hint_raw=hint_raw,
        ocr_score=ocr_score,
        source_sheet=ws.title,
        source_row=row_idx,
    )


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _file_md5(path: Path) -> str:
    return hashlib.md5(path.read_bytes()).hexdigest()


def mark_suspicious_placeholder_pet_icons(pets: list[dict[str, Any]], foods: list[dict[str, Any]]) -> None:
    pet_hash_to_indices: dict[str, list[int]] = {}
    food_hashes: set[str] = set()

    for index, food in enumerate(foods):
        if food.get("iconMissing"):
            continue
        alias = food.get("iconAliasRel")
        if not alias:
            continue
        icon_path = ROOT_DIR / alias
        if not icon_path.exists():
            continue
        food_hashes.add(_file_md5(icon_path))

    for index, pet in enumerate(pets):
        if pet.get("iconMissing"):
            continue
        alias = pet.get("iconAliasRel")
        if not alias:
            continue
        icon_path = ROOT_DIR / alias
        if not icon_path.exists():
            continue
        h = _file_md5(icon_path)
        pet_hash_to_indices.setdefault(h, []).append(index)

    overlapping_hashes = set(pet_hash_to_indices).intersection(food_hashes)
    for h in overlapping_hashes:
        for pet_index in pet_hash_to_indices[h]:
            pet = pets[pet_index]
            if pet.get("implStatus") == "placeholder":
                pet["iconMissing"] = True


def build_catalog() -> dict[str, Any]:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Missing source config: {SOURCE_XLSX}")
    if not ICONS_DIR.exists():
        raise FileNotFoundError(f"Missing icon source dir: {ICONS_DIR}")

    wb = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    sheet_names = wb.sheetnames[:6]
    if len(sheet_names) < 6:
        raise RuntimeError("Expected 6 pack sheets in configs.xlsx")

    dedup: dict[tuple[str, str, str], RawRow] = {}
    for sheet in sheet_names:
        ws = wb[sheet]
        for row_idx in range(2, ws.max_row + 1):
            parsed = parse_row(ws, row_idx)
            if not parsed:
                continue
            key = (parsed.pack_name_zh, parsed.type_zh, normalize_name(parsed.name_raw))
            prev = dedup.get(key)
            if not prev or parsed.ocr_score > prev.ocr_score:
                dedup[key] = parsed

    rows = sorted(
        dedup.values(),
        key=lambda r: (
            PACK_KEY_MAP[r.pack_name_zh]["key"],
            0 if r.type_zh == "动物" else 1,
            r.tier,
            normalize_name(r.name_raw),
            r.source_row,
        ),
    )
    if not rows:
        raise RuntimeError("No valid rows parsed from configs.xlsx")

    if ICONS_EN_DIR.exists():
        shutil.rmtree(ICONS_EN_DIR)
    ICONS_EN_DIR.mkdir(parents=True, exist_ok=True)

    seq_by_pack_type: dict[tuple[str, str], int] = {}
    packs: list[dict[str, Any]] = []
    pack_seen: set[str] = set()
    pets: list[dict[str, Any]] = []
    foods: list[dict[str, Any]] = []
    catalog_rows: list[dict[str, Any]] = []

    for raw in rows:
        pack_meta = PACK_KEY_MAP[raw.pack_name_zh]
        pack_key = pack_meta["key"]
        if pack_key not in pack_seen:
            pack_seen.add(pack_key)
            packs.append(
                {
                    "key": pack_key,
                    "nameZh": raw.pack_name_zh,
                    "nameEn": pack_meta["name_en"],
                }
            )

        type_key = "pet" if raw.type_zh == "动物" else "food"
        seq_key = (pack_key, type_key)
        next_seq = seq_by_pack_type.get(seq_key, 0) + 1
        seq_by_pack_type[seq_key] = next_seq
        item_id = f"{pack_key}_{type_key}_{next_seq:04d}"

        normalized = normalize_name(raw.name_raw)
        hint_clean = clean_hint(raw.hint_raw)

        icon_src_rel = raw.icon_src_rel.lstrip("./")
        src_path = ICONS_DIR / icon_src_rel
        alias_rel = f"assets/icons_en/{pack_key}/{type_key}s/{item_id}.png"
        alias_path = ROOT_DIR / alias_rel
        icon_missing = not src_path.exists()
        if src_path.exists():
            ensure_parent(alias_path)
            shutil.copyfile(src_path, alias_path)

        if type_key == "pet":
            legacy = LEGACY_PET_KIND_MAP.get(normalized)
            if legacy:
                kind = legacy["kind"]
                name_en = legacy["name_en"]
                ability_key = legacy["ability_key"]
                attack = int(legacy["attack"])
                health = int(legacy["health"])
                impl_status = "implemented"
                color = legacy["color"]
            else:
                kind = item_id
                name_en = ""
                ability_key = "placeholder_none"
                attack, health = auto_pet_stats(item_id, raw.tier)
                impl_status = "placeholder"
                color = color_from_id(item_id, is_food=False)

            pet = {
                "id": item_id,
                "kind": kind,
                "packKey": pack_key,
                "nameZh": normalized,
                "nameEn": name_en,
                "tier": raw.tier,
                "roundUnlock": raw.round_unlock,
                "iconSrcRel": icon_src_rel,
                "iconAliasRel": alias_rel,
                "iconMissing": icon_missing,
                "hintZhRaw": raw.hint_raw,
                "hintZhClean": hint_clean,
                "hintEn": "",
                "attack": attack,
                "health": health,
                "abilityKey": ability_key,
                "implStatus": impl_status,
                "color": color,
                "sourceSheet": raw.source_sheet,
                "sourceRow": raw.source_row,
            }
            pets.append(pet)
            catalog_rows.append(
                {
                    "id": item_id,
                    "pack_key": pack_key,
                    "pack_name_zh": raw.pack_name_zh,
                    "pack_name_en": pack_meta["name_en"],
                    "type": "pet",
                    "name_zh": normalized,
                    "name_en": name_en,
                    "tier": raw.tier,
                    "round_unlock": raw.round_unlock,
                    "icon_src_rel": icon_src_rel,
                    "icon_alias_rel": alias_rel,
                    "hint_zh_raw": raw.hint_raw,
                    "hint_zh_clean": hint_clean,
                    "hint_en": "",
                    "attack": attack,
                    "health": health,
                    "ability_key": ability_key,
                    "effect_key": "",
                    "impl_status": impl_status,
                    "source_sheet": raw.source_sheet,
                    "source_row": raw.source_row,
                }
            )
        else:
            legacy = LEGACY_FOOD_KIND_MAP.get(normalized)
            if legacy:
                kind = legacy["kind"]
                name_en = legacy["name_en"]
                effect_key = legacy["effect_key"]
                impl_status = "implemented"
                color = legacy["color"]
            else:
                kind = item_id
                name_en = ""
                effect_key = "placeholder_stat_buff"
                impl_status = "placeholder"
                color = color_from_id(item_id, is_food=True)

            buff = placeholder_food_buff(raw.tier)
            food = {
                "id": item_id,
                "kind": kind,
                "packKey": pack_key,
                "nameZh": normalized,
                "nameEn": name_en,
                "tier": raw.tier,
                "roundUnlock": raw.round_unlock,
                "iconSrcRel": icon_src_rel,
                "iconAliasRel": alias_rel,
                "iconMissing": icon_missing,
                "hintZhRaw": raw.hint_raw,
                "hintZhClean": hint_clean,
                "hintEn": "",
                "effectKey": effect_key,
                "implStatus": impl_status,
                "placeholderBuff": buff,
                "color": color,
                "sourceSheet": raw.source_sheet,
                "sourceRow": raw.source_row,
            }
            foods.append(food)
            catalog_rows.append(
                {
                    "id": item_id,
                    "pack_key": pack_key,
                    "pack_name_zh": raw.pack_name_zh,
                    "pack_name_en": pack_meta["name_en"],
                    "type": "food",
                    "name_zh": normalized,
                    "name_en": name_en,
                    "tier": raw.tier,
                    "round_unlock": raw.round_unlock,
                    "icon_src_rel": icon_src_rel,
                    "icon_alias_rel": alias_rel,
                    "hint_zh_raw": raw.hint_raw,
                    "hint_zh_clean": hint_clean,
                    "hint_en": "",
                    "attack": "",
                    "health": "",
                    "ability_key": "",
                    "effect_key": effect_key,
                    "impl_status": impl_status,
                    "source_sheet": raw.source_sheet,
                    "source_row": raw.source_row,
                }
            )

    packs.sort(key=lambda item: item["key"])
    pet_by_kind: dict[str, dict[str, Any]] = {}
    for pet in pets:
        pet_by_kind.setdefault(pet["kind"], pet)
    food_by_kind: dict[str, dict[str, Any]] = {}
    for food in foods:
        food_by_kind.setdefault(food["kind"], food)

    mark_suspicious_placeholder_pet_icons(pets, foods)

    payload = {
        "version": datetime.now(timezone.utc).isoformat(),
        "packs": packs,
        "pets": pets,
        "foods": foods,
        "indexes": {
            "petByKind": pet_by_kind,
            "foodByKind": food_by_kind,
        },
    }
    return {"payload": payload, "catalog_rows": catalog_rows}


def write_output_xlsx(catalog_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "catalog"
    ws.append(CATALOG_COLUMNS)
    for row in catalog_rows:
        ws.append([row.get(col, "") for col in CATALOG_COLUMNS])
    wb.save(OUTPUT_XLSX)


def write_output_js(payload: dict[str, Any]) -> None:
    OUTPUT_JS.parent.mkdir(parents=True, exist_ok=True)
    data_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    content = "// Auto-generated by scripts/build_game_catalog.py\nwindow.__GAME_DATA = " + data_json + ";\n"
    OUTPUT_JS.write_text(content, encoding="utf-8")


def main() -> int:
    built = build_catalog()
    payload = built["payload"]
    catalog_rows = built["catalog_rows"]
    write_output_xlsx(catalog_rows)
    write_output_js(payload)

    print(f"Generated: {OUTPUT_XLSX}")
    print(f"Generated: {OUTPUT_JS}")
    print(f"Generated alias icons under: {ICONS_EN_DIR}")
    print(f"packs={len(payload['packs'])} pets={len(payload['pets'])} foods={len(payload['foods'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
